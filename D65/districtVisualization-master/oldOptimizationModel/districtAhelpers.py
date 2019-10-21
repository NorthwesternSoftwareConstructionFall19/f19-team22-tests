# -*- coding: utf-8 -*-
"""
Districting helper functions
Dec 2018/Jan 2019

Northwestern/D65 Team

BR and MLK included but magnet program location fixed
"""

import pandas as pd
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
from gurobipy import *


# read school capacities from file, given number of schools N
def readcapacities(capacityfile, sheet, N):
    print("\nReading school capacities...\n")
    
    xl = pd.ExcelFile(capacityfile)
    caps = xl.parse(sheet)
    
    c = {}
    
    for n in range(N):
        c[n] = caps.loc[n, 0]
    
    return c


# define program eligibilities, for now m == n
def eligible(M, N):
    print("Setting program eligibilities...\n")
    
    gamma = {}
    for m in range(M):
        for n in range(N):
            if m == n:
                gamma[m, n] = 1
            else:
                gamma[m, n] = 0
            
    return gamma


# create I_set  - set of blocks that are in the district
# create l[i,n] - binary matrix indicating which blocks are assigned to which schools
def blocksetup(currentassignmentsfile, I, N):
    print("Setting up blocks...\n")
    
    file = load_workbook(currentassignmentsfile)
    main = file["main"]
    
    I_set = []
    l = {}
    
    for i in range(I):
        if main.cell(i + 2, 5).value < 100:
            I_set.append(i)
            for n in range(N):
   #deleted row and column notation in cell method
                if main.cell(i + 2, 5).value == n:
                    l[i, n] = 1
                else:
                    l[i, n] = 0
                
    return I_set, l


# create W_pre - dictionary of adjacency sets
#deleted row and column notation in cell method
def blockadjacency(adjfile, I_set):
    print("Loading block adjacency file...\n")
        
    file = load_workbook(adjfile)
    main = file["main"]
       
    print("Reading block adjacencies...\n")
    W_pre = {}
    
    for i in I_set:
        W_pre[i] = []
        
        for j in I_set:
            if (main.cell(i + 2, j + 2).value == 1) and (i != j):
                W_pre[i].append(j)
                
    return W_pre


# create W - where W[i, n] == set of blocks adjacent to block i which are closer to school n than block i is
#deleted row and column notation in cell method
def blockcontiguity(distfile, I_set, N, W_pre):
    print("Loading Euclidean distances file...\n")
        
    file = load_workbook(distfile)
    main = file["main"]
    
    print("Reading Euclidean distances...\n")
    
    dist = {}
    
    curr_row = 2
    
    while main.cell(curr_row, 1).value is not None:
        i = main.cell(curr_row, 1).value
        n = main.cell(curr_row, 2).value
        d = main.cell(curr_row, 4).value
        
        dist[i, n] = d
        
        curr_row = curr_row + 1
        
    print("Setting up dictionary for contiguity constraints...\n")
    
    W = {}
    
    for i in I_set:
        for n in range(N):
            W[i, n] = []
            for j in W_pre[i]:
                if dist[j, n] <= dist[i, n]:
                    W[i, n].append(j)
                    
    return(W)


# create d - where d[i, n] == true distance from block i to school n
#deleted row and column notation in cell method
def truedistances(distfile, I_set, N):
    print("Loading true (walking) distances file...\n")
    
    file = load_workbook(distfile)
    main = file["main"]

    print("Reading true distances...\n")       

    dist = {}
    
    curr_row = 2
    
    while main.cell(curr_row, 1).value is not None:
        i = main.cell(curr_row, 1).value
        if i in I_set:
            n = main.cell(curr_row, 2).value
            d = main.cell(curr_row, 4).value
            
            dist[i, n] = d
        
        curr_row = curr_row + 1
        
    return dist

def busblocks(busfile, I_set, N):
    print("Loading busing data...")
    
    bus = {}
    bustrue = {}
    
    file = load_workbook(busfile)
    main = file["main"] 
    
    curr_row = 2
    while main.cell(curr_row, 1).value is not None:
        i = main.cell(curr_row, 1).value

        if i in I_set:
            for n in range(N):
                curr_col = n + 2
                bus[i, n] = main.cell(curr_row, curr_col).value
                bustrue[i, n] = bus[i, n]
                
                if bus[i, n] == 2:
                    bus[i, n] = 1
                    
        curr_row = curr_row + 1
        
    return bus, bustrue
        

        
# create h - where h[i, n] == 1 if block i contains school n, 0 otherwise
def schoolinblock(I_set, N):
    h = {}
    for i in I_set:
        for n in range(N):
            h[i, n] = 0
            
    h[519, 0] = 1 #
    h[538, 1] = 1 #
    h[375, 2] = 1 #
    h[641, 3] = 1 #
    h[224, 4] = 1 #
    h[620, 5] = 1 #
    h[748, 6] = 1 #
    h[113, 7] = 1 #
    h[230, 8] = 1 #
    h[31, 9] = 1 #
    h[112, 10] = 1 #
    h[485, 11] = 1 #
    
    return h


# create population parameters from file
def popsetup(popfile, I_set, M, N, J, S, T, l):
    print("Loading population data file...\n")
    
    file = load_workbook(popfile)
    main = file["main"]

    print("Reading population data...\n")       

    p = {}
    q = {}
    p_exp = {}
    q_exp = {}
    
    for i in I_set:
        for j in range(J):
            p_exp[i, j] = 0
            
            for m in range(M):
                q_exp[i, j, m] = 0
            
                for s in range(S):
                    for t in range(T):
                        q[i, j, m, s, t] = 0
                    
    
    curr_row = 2
    
    while main.cell(curr_row, 1).value is not None:
        i = main.cell(curr_row, 1).value
        
        if i in I_set:
            s = main.cell(curr_row, 2).value
            t = main.cell(curr_row, 3).value
            
            for n in range(N):
                if l[i, n] == 1:
                    for j in range(J):
                        jcol = (n + 2) * 2 + j
                        p[i, j, s, t] = main.cell(curr_row, jcol).value
                else:
                    m = n
                    for j in range(J):
                        jcol = (m + 2) * 2 + j                        
                        q[i, j, m, s, t] = q[i, j, m, s, t] + main.cell(curr_row, jcol).value

        curr_row = curr_row + 1
        
    for i in I_set:
        for j in range(J):
            for s in range(S):
                for t in range(T):
                    p_exp[i, j] = p_exp[i, j] + p[i, j, s, t]
                
            p_exp[i, j] = p_exp[i, j]/(S * T)
        
            
            for m in range(M):
                q_exp[i, j, m] = 0
                for s in range(S):
                    for t in range(T):
                        q_exp[i, j, m] = q_exp[i, j, m] + q[i, j, m, s, t]
                    
                    q_exp[i, j, m] = q_exp[i, j, m]/(S * T)
        
            # q_exp_tot[m] = q_exp_tot[m] + q_exp[i, m]
            
    return p, q, p_exp, q_exp

'''
# new function to read population/enrollment data in the form of demo.xlsx
### NEEDS FIXING ONCE NEW DATA DELIVERED ###
def popsetup(popfile, I_set, M, N, N, S, T, l):
    print("Loading population data file...\n")
    
    file = load_workbook(popfile)
    main = file["main"]

    print("Reading population data...\n")       

    p = {}
    q = {}
    p_exp = {}
    q_exp = {}
    
    for i in I_set:
        p_exp[i] = 0
            
        for m in range(M):
            q_exp[i, m] = 0
            
            for s in range(S):
                for t in range(T):
                    q[i, m, s, t] = 0
                    
    
    curr_row = 2
    
    while main.cell(row = curr_row, column = 1).value is not None:
        i = main.cell(row = curr_row, column = 1).value
        
        if i in I_set:
            s = main.cell(row = curr_row, column = 2).value
            t = main.cell(row = curr_row, column = 3).value
            
            for n in range(N):
                if l[i, n] == 1:
                    p[i, s, t] = main.cell(row=curr_row, column= jcol).value
                else:
                    m = n
                    for j in range(J):
                        jcol = (m + 2) * 2 + j                        
                        q[i, j, m, s, t] = q[i, j, m, s, t] + main.cell(row=curr_row, column= jcol).value

        curr_row = curr_row + 1
        
    for i in I_set:
        for j in range(J):
            for s in range(S):
                for t in range(T):
                    p_exp[i, j] = p_exp[i, j] + p[i, j, s, t]
                
            p_exp[i, j] = p_exp[i, j]/(S * T)
        
            
            for m in range(M):
                q_exp[i, j, m] = 0
                for s in range(S):
                    for t in range(T):
                        q_exp[i, j, m] = q_exp[i, j, m] + q[i, j, m, s, t]
                    
                    q_exp[i, j, m] = q_exp[i, j, m]/(S * T)
        
            # q_exp_tot[m] = q_exp_tot[m] + q_exp[i, m]
            
    return p, q, p_exp, q_exp


'''

# blockchars reads in the demographic data
def blockchars(charfile, I_set, J, R, TP):
    print("Loading population characteristic data")
    
    # load in the demographic data file and set page to main
    file = load_workbook(charfile)
    main = file["main"]
    
    # create rac[i,r] which records how many students from block i are of race r
    rac = {}
    # create twi[i] which records how many students from block i are in TWI
    twi = {}
    # create frl[i] which records how many students from block i are FRL eligible
    frl = {}
    # create grades[i, j]
    grades = {}
    for i in I_set:
        twi[i] = 0
        frl[i] = 0
        for r in range(R):
            rac[i,r] = 0
        for j in range(J):
            grades[i, j] = 0
    
    # set current row to 2 to initialize reading data from demographic file
    curr_row = 2
    
    # while loop will run until the first cell in a row is empty
    while main.cell(row = curr_row, column = 1).value is not None:
        i = main.cell(row = curr_row, column = 1).value
        # for each block i, and each grade bucket j, grades[i, j] stores how many students there are
        # grade bucket breakdowns are stored in columns 18, 19, and 20
        for j in range(J):
            curr_col = j + 18
            grades[i, j] = main.cell(row = curr_row, column = curr_col).value
        
        # for each block i, twi[i] stores the sum of students in each type of TWI program
        # TWI program enrollments per block are stored in columns 21, 22, and 23 of demo.xlsx
        twi[i] = main.cell(row = curr_row, column = 21).value + main.cell(row = curr_row, column = 22).value + main.cell(row = curr_row, column = 23).value            
        
        # for each block i and race group r, rac[i, r] stores the number of students
        # race data is stored in columns 24 - 28
        for r in range(R):
            curr_col = r + 24
            rac[i, r] = main.cell(row = curr_row, column = curr_col).value 
        
        # for each block i, frl[i] stores the number of FRL-eligible students
        # FRL data is stored in column 29
        frl[i] = main.cell(row = curr_row, column = 29).value
                    
        curr_row = curr_row + 1
    
    return twi, rac, frl, grades

# feederschools reads the feeder matrix and creates the matrix f
def feederschools(feederfile, N, K):
    print("Loading feeder school data file...\n")
    
    # load in the feeders data file and set page main
    file = load_workbook(feederfile)
    main = file["main"]
    
    # f[i,j] is a binary variable = 1 if school i feeds into school j, 0 otherwise
    f = {}
    
    # double for loop to initialize the 2-dimensional f array to store feeder indicators
    for n in range(N):
        for k in range(K):
            f[n, k] = 0
    
    # set current row to be 2; the first row of actual data in the table        
    curr_row = 2
    
    # while loop will run until the first cell in a row is empty
    while main.cell(row = curr_row, column = 1).value is not None:
        n = main.cell(row = curr_row, column = 1).value
        
        for k in range(K):
            curr_col = k + 2
            f[n, k] = main.cell(row = curr_row, column = curr_col).value
                    
        curr_row = curr_row + 1
    
    return f

def buscalcs(resultsfile, I_set, M, N, J, blocks, progs, p_exp, q_exp, bustrue, busperschool):
    
    distbus = {}
    hazbus= {}
    
    for i in I_set:
        distbus[i] = 0
        hazbus[i] = 0
        
        for n in range(N):            
            if blocks[i] == n:    
                if bustrue[i, n] == 1:
                    for j in range(J):
                        hazbus[i] = hazbus[i] + p_exp[i, j]
                elif bustrue[i, n] == 2:
                    for j in range(J):
                        distbus[i] = distbus[i] + p_exp[i, j] 
                    
            for m in range(M):
                if progs[m] == n:
                    if bustrue[i, n] == 1:
                        for j in range(J):
                            hazbus[i] = hazbus[i] + q_exp[i, j, m]
                    elif bustrue[i, n] == 2:
                        for j in range(J):
                            distbus[i] = distbus[i] + q_exp[i, j, m]



    file = load_workbook(filename = resultsfile)
    main = file["busing"]
    
    curr_row = 2
    
    for i in I_set:
        main.cell(curr_row, 1).value = i
        main.cell(curr_row, 2).value = distbus[i]
        main.cell(curr_row, 3).value = hazbus[i]
        
        curr_row = curr_row + 1
        
    curr_row = 2
    
    for n in range(N):
        main.cell(curr_row, 5).value = n
        main.cell(curr_row, 6).value = busperschool[n]
        
        curr_row = curr_row + 1

    file.save(resultsfile)
        
        
def writeresults(resultsfile, I_set, M, N, J, S, T, R, capacities, blocks,
                 progs, enrollments, overcaps, exptrav, contig_flag, sch_twi,
                 sch_rac, sch_frl, sch_grades, totnumbus, totpcbus, t1elig):
    print("Writing results...\n")
    
    file = load_workbook(resultsfile)
    
    # print assignment summary to sheet "main"
    main = file["main"]

    for i in I_set:
        main.cell(i + 2, 6).value = blocks[i]
        
    if contig_flag == 1:
        main.cell(3, 12).value = 'YES'
    else:
        main.cell(3, 12).value = 'NO'
       
    main.cell(4, 12).value = exptrav/1.609 # convert to miles
    
    totovercap = 0
    for n in range(N):
        totovercap = totovercap + overcaps[n]
        
    main.cell(5, 12).value = totovercap/N
    main.cell(6, 12).value = totnumbus
    main.cell(7, 12).value = 100 * totpcbus
    
    
    
    # print list of programs to sheet "programs"
    programs = file["programs"]    
    for m in range(M):
        programs.cell(m + 2, 4).value = progs[m]
    
    # print list of schools and projected enrollments to sheet "enrollments"
    enr = file["enrollments"]
    
    curr_row = 2
   
    for n in range(N):
        for s in range(S):
            for t in range(T):
                for j in range(J):
                    enr.cell(curr_row, 1).value = n
                    enr.cell(curr_row, 2).value = s
                    enr.cell(curr_row, 3).value = t
                    enr.cell(curr_row, 4).value = j
                    enr.cell(curr_row, 5).value = capacities[n]
                    enr.cell(curr_row, 6).value = enrollments[n, j, s, t]
    
                    curr_row = curr_row + 1
    
    # print school-level demographic data in sheet "demographics"
    demo = file["demographics"]
    
    curr_row = 2
    
    for n in range(N):
        demo.cell(curr_row, 1).value = n
        demo.cell(curr_row, 2).value = sch_twi[n]
        demo.cell(curr_row, 3).value = sch_frl[n]
        for r in range(R):
            curr_col = 4
            demo.cell(curr_row, curr_col + r).value = sch_rac[n, r]
        for j in range(J):
            curr_col = 9
            demo.cell(curr_row, curr_col + j).value = sch_grades[n, j]
        demo.cell(curr_row, 11).value = t1elig[n]
            
        curr_row = curr_row + 1


    file.save(resultsfile)
    

        
        
        
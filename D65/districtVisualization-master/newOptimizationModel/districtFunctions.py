# -*- coding: utf-8 -*-
"""
New districting helper functions
Created on Thu Jul 11 09:00:38 2019

@author: f3lix
"""
# import necessary packages
import pandas as pd
import numpy as np
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
from gurobipy import *

###------------------------------------------------------------------------###
############################### FUNCTIONS ####################################
###------------------------------------------------------------------------###

# Function 1: blocksetup: read number of blocks and current assignments
def blocksetup(blockfile, N):
    print("Setting up blocks...\n")
    temp = pd.read_csv(blockfile, index_col = None, header = None)              # read in file, then convert to an array
    temp = temp.to_numpy()
    
    I = len(temp)                                                               # save number of blocks as I
    I_set = []                                                                  # create an empty set I_set to which we will add all in-district blocks
    for i in range(I):
        if temp[i] < 100:
            I_set.append(i)                                                     # if a block is not assigned to 300, then it is in-district
    
    l = np.zeros((len(I_set), N))                                               # create an IxN array l which is 1 if block i currently assigned to school n
    for i in range(len(I_set)):
        n = temp[I_set[i]]                                                      # reads assignment of block i for each i in the in-district blocks in I_set
        l[i, n] = 1                                                             # updates entry in l to 1 for (i, n) pair
    
    return I, I_set, l

# Function 2: capacities: read school capacities
def capacities(capfile):
    print("Setting up school capacities...\n")
    temp = pd.read_csv(capfile, index_col = None, header = None, dtype = int)   # read capacities file, then convert to array
    
    c = temp.to_numpy()
    c = np.concatenate(c, axis = 0)                                             # transforms array of arrays into a 1D array
    
    return c

# Function 3: schools: read school-block adjacency relationships
def schools(I, I_set, N):
    print("Setting up school locations...\n")
    s0 = [122, 492, 515, 517, 518]
    s1 = [534, 535, 536, 537, 538, 539, 540, 541, 542, 588, 591, 709]
    s2 = [75, 374, 383, 390, 391, 527, 528, 530]
    s3 = [639, 640, 666, 667, 668, 641, 643, 644, 645]
    s4 = [187, 190, 196, 223, 401, 404, 405, 407]
    s5 = [616, 617, 619, 620, 622, 630, 788, 789]
    s6 = [729, 730, 732, 746, 747]
    s7 = [109, 112, 113, 115, 119, 132, 137, 326]
    s8 = [51, 59, 60, 224, 225, 229, 456]
    s9 = [25, 29, 30, 163, 370, 371, 373]
    s10 = []                                                                    # magnet school, enrollment will not depend on distance
    s11 = []                                                                    # magnet school, enrollment will not depend on distance
    temp = [s0, s1, s2, s3, s4, s5, s6, s7, s8, s9, s10, s11]                   # create array of arrays
    
    hp = np.zeros((I, N), dtype = int)
    for i in range(len(temp)):
        for j in range(len(temp[i])):
            ind1 = temp[i][j]                                                   # ind1 is the block number (indexing begins at 0 for block 1)
            ind2 = i                                                            # ind2 is the school number
            hp[ind1, ind2] = 1
    
    h = np.zeros((len(I_set), N), dtype = int)                                  # transform h to only include relevant blocks
    for i in range(len(I_set)):
        for n in range(N):
            ind = I_set[i]
            h[i, n] = hp[ind, n]
            
    return h

# Function 4: adjacency: read segments and designate block adjacencies
def adjacency(segfile, bridgefile, islandfile, I, I_set):
    print("Setting up block adjacencies...\n")
    temp = pd.read_csv(segfile, index_col = None, header = None, dtype = int)   # read in segment-block data and convert to numpy array
    
    segs = temp.to_numpy()
    nseg = segs[:, 0].max()                                                     # determine number of segments as maximum of the column of segment IDs
    segs = segs[:, 1].reshape((nseg, 2))                                        # transform array s.t. each row contains the blocks that a segment borders
    
    adj = np.zeros((I, I), dtype = int)                                         # initialize edge-adjacency matrix, 1 if blocks i1 and i2 share an edge
    for i in range(nseg):
        ind1 = segs[i, 0]
        ind2 = segs[i, 1]
        adj[ind1, ind2] = 1
        adj[ind2, ind1] = 1                                                     # set pair of blocks and their reverse to 1 in the adjacency matrix
    
    for i in range(I):
        adj[i, i] = 0                                                           # Removing block adjacency relationships with self
    
    print("Adding bridge adjacencies...\n")
    temp = pd.read_csv(bridgefile, index_col = None, header = None, dtype = int)

    bridges = temp.to_numpy()
    nbridge = len(bridges)
    for i in range(nbridge):
        ind1 = bridges[i, 0]
        ind2 = bridges[i, 1]
        adj[ind1, ind2] = 1
        adj[ind2, ind1] = 1
    
    print("Adding manual adjacencies...\n")
    temp = pd.read_csv(islandfile, index_col = None, header = None, dtype = int)
    
    islands = temp.to_numpy()
    nisland = len(islands)
    for i in range(nisland):
        ind1 = islands[i, 0]
        ind2 = islands[i, 1]
        adj[ind1, ind2] = 1
        adj[ind2, ind1] = 1
    
    adj = adj[I_set, :]                                                         # cutting adj down to only include blocks in I_set
    adj = adj[:, I_set]
    
    return adj

# Function 5: contiguity: build contiguity sets based on distances to schools
def contiguity(distfile, adj, I, I_set, N):
    print("Building contiguity relationships...\n")
    temp = pd.read_csv(distfile, index_col = None, header = None)               # read in euclidean distances from blocks to schools
    
    u = temp.to_numpy()
    u = u.reshape((I, N))                                                       # reshape array into IxN, shows euclidean distance from block i to school n
    u = u[I_set, :]
    
    II = len(I_set)
    
    W = []                                                                     
    for i in range(II):
        for n in range(N):
            lst = []
            for j in range(II):
                if adj[i, j] == 1:
                    if u[j, n] <= u[i, n]:
                        lst.append(j)
            W.append(lst)
    
    W = np.reshape(W, (II, N))
    
    return W

# Function 6: truedist: read true (walking/busing) distances
def truedist(truedistfile, E, I, I_set, N, h):
    print("Setting up distances...\n")
    temp = pd.read_csv(truedistfile, index_col = None, header = None)
    
    d = temp.to_numpy()
    d = d.reshape((I, N))                                                       # transform array into IxN, display distance from block i to school n
    
    d = d[I_set, :]
    
    for i in range(len(I_set)):                                                 # based on busing distance, set radius of blocks within distance E of n
        for n in range(N):                                                      # to automatically be assigned to school n
            if d[i, n] <= E:
                h[i, n] = 1
    
    return h, d

# Function 7: busreqs: read hazard/busing matrix
def busreqs(busfile, I_set):
    print("Reading block-to-school busing requirements...\n")
    temp = pd.read_csv(busfile, index_col = None, header = None, dtype = int)
    
    b = temp.to_numpy()
    b = b[I_set, :]
    
    return b

# Function 8: enrollments: read enrollment data
def enrollments(enrollfile, I_set, J, S, T, N):
    print("Setting up enrollments...\n")
    temp = pd.read_csv(enrollfile, index_col = None, header = None, dtype = int)
    
    e = temp.to_numpy()
    e = e.reshape((len(I_set), J, S, T, N))                                     # using the I_set instead of I; all enrollment #s will only use relevant i's
    
    q = e[..., 10:11]                                                           # isolate the enrollments into magnet schools, store in array q
    q = q.sum(axis = 4)                                                         # sum along schools to get per-grade population of block i in scen s, year t
    p = e
    p[..., 10:11] = 0                                                           # remove magnet school enrollments, save new array as p
    p = p.sum(axis = 4)                                                         # sum along schools to get per grade population of block i in scen s, year t
    
    qbar = q.sum(axis = (1, 2, 3))/(S * T)                                      # qbar is the average # students in i attending magnet schools
    pbar = p.sum(axis = (1, 2, 3))/(S * T)                                      # pbar is the average # students in i attending neighborhood schools
    
    phi = 0
    rho = 15
    psi = 0
    
    return e, p, q, pbar, qbar, phi, rho, psi

# Function 9: demographics: read population characteristics/demographics
def demographics():
    print("Seting up population demographics...\n")


# Function 10: busing: calculate and write post-assignment busing situation
def busing():
    print("Calculating busing results...\n")

# Function 11: writeresults: write results of assignment
def writeresults(resultfile, assn, cap, I_set, b, d, pbar, qbar):
    print("Writing results...\n")
    file = load_workbook(resultfile)                                            # load results workbook
    enroll = file['enroll']                                                     # load sheet 1 to record enrollment results
    bus = file['bus']                                                           # load sheet 2 to record busing results
    dist = file['dist']                                                         # load sheet 3 to record distance results
    
    II = len(I_set)
    for i in range(II):
        enroll.cell(row = i + 2, column = 1).value = I_set[i] + 1
        enroll.cell(row = i + 2, column = 6).value = assn[i]                    # record assignment of block i in column F for reading into visualization
        enroll.cell(row = i + 2, column = 7).value = pbar[i]                    # record average population of block i into column G for calculations
        
        bus.cell(row = i + 2, column = 1).value = I_set[i] + 1
        bus.cell(row = i + 2, column = 2).value = assn[i]
        bus.cell(row = i + 2, column = 3).value = b[i, assn[i]]                 # record whether or not block i needs to be bused to its assigned school
        bus.cell(row = i + 2, column = 4).value = pbar[i]                       # record average population of block i for calculations
        
        dist.cell(row = i + 2, column = 1).value = I_set[i] + 1
        dist.cell(row = i + 2, column = 2).value = assn[i]
        dist.cell(row = i + 2, column = 3).value = d[i, assn[i]]                # record distance from block i to its assigned school
        
    file.save(resultfile)
    print("Program complete\n")

###------------------------------------------------------------------------###
############################# SUB-FUNCTIONS ##################################
###------------------------------------------------------------------------###

# Function A1: contig(): checks if the adjacency matrix represents a connected network
def contig(matrix, block, path=[]):
    blocks = len(matrix)
    path.append(block)
    
    for i in range(blocks):
        if matrix[block, i] == 1:
            if i not in path:
                path = contig(matrix, i, path)                                  # recursively build a path of blocks connected to start via adjacency
    
    return path

# Function A2: find_missing(): checks the output of contig() to see which blocks are missing
def find_missing(lst):
    lst = sorted(lst)                                                           # numerically sorts list
    miss = [x for x in range(lst[0], lst[-1]+1) if x not in lst]                # assuming numerical order, finds gaps in the list
    
    return miss

